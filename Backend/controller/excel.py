from openpyxl import load_workbook
from openpyxl import Workbook
from datetime import datetime
import json


class excel_to_json():
    def __init__(self, filename):
        self.workbook = load_workbook(f'{filename}')
        self.data = {}
        self.f_json = []
        self.process_data()

    def process_data(self):
        sheet_name = self.workbook.sheetnames[0]
        worksheet = self.workbook[sheet_name]
        for i, row in enumerate(worksheet.iter_rows()):
            if i == 0:
                title = [cell.value for cell in row]
            else:
                for i, cell in enumerate(row):
                    self.data[title[i]] = cell.value
                self.f_json.append(self.data)
                self.data = {}

    def return_json(self):
        print (self.f_json)
        return json.dumps(self.f_json)
